import openpyxl
import requests
import os
from dotenv import load_dotenv
import openai

load_dotenv()

api_key = os.getenv("OPENAI_KEY")
openai.api_key = api_key

def chat_with_gpt3(prompt):
    try:
        response = openai.Completion.create(
            engine="text-davinci-002",  # The GPT-3.5 chat model
            prompt= prompt + " with the response being provided in a poetic style",  # Add your string to the prompt
            max_tokens=150  # Adjust this value to control the response length
        )
        print(response.choices[0].text.strip())
        return response.choices[0].text.strip()
    except Exception as e:
        return str(e)


def process_excel_file(file_name):
    current_directory = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(current_directory, file_name)
    
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    
    # Add 'Completions' to the first row of the second column
    sheet.cell(row=1, column=2, value='Completions')

    for idx, row in enumerate(sheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True), start=2):
        prompt = row[0]
        completion = chat_with_gpt3(prompt)
        sheet.cell(row=idx, column=2, value=completion)


    workbook.save(file_path)


file_name = 'example_prompt_data.xlsx'


process_excel_file(file_name)
